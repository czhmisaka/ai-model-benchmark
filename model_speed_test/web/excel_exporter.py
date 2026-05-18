"""
Excel 报告导出服务
用于导出多 Sheet 的 Excel 报告
"""
import io
from typing import Dict, Any, List, Optional
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelExporter:
    """Excel 报告导出器"""
    
    # 样式定义
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl 未安装。请运行: pip install openpyxl>=3.1.0"
            )
    
    def export(self, data: Dict[str, Any]) -> bytes:
        """
        导出多 Sheet Excel
        
        Args:
            data: 包含 summary, results, model_stats, chart_data 的字典
            
        Returns:
            Excel 文件的字节数据
        """
        wb = Workbook()
        
        # 移除默认 sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Sheet 1: 汇总统计
        self._create_summary_sheet(wb, data.get('summary', {}))
        
        # Sheet 2: 模型对比
        self._create_model_comparison_sheet(wb, data.get('model_stats', []))
        
        # Sheet 3: 详细数据
        self._create_detail_sheet(wb, data.get('results', []))
        
        # Sheet 4: 图表数据（如果有）
        chart_data = data.get('chart_data', {})
        if chart_data:
            self._create_chart_data_sheet(wb, chart_data)
        
        # 保存到 buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    def _create_summary_sheet(self, wb: Workbook, summary: Dict[str, Any]):
        """
        汇总统计 Sheet
        
        Args:
            wb: Workbook 实例
            summary: 汇总统计数据
        """
        ws = wb.create_sheet("汇总统计")
        
        # 标题
        ws['A1'] = '模型速度测试 - 汇总统计'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # 副标题
        ws['A2'] = f"测试组 ID: {summary.get('group_id', '')}"
        ws['A2'].font = Font(size=11, italic=True)
        
        # 统计数据
        stats = [
            ('总测试数', summary.get('total', 0)),
            ('成功数', summary.get('success', 0)),
            ('失败数', summary.get('failed', 0)),
            ('成功率', f"{summary.get('success_rate', 0):.1f}%"),
            ('平均 TTFT', f"{summary.get('avg_ttft', 0):.3f}s ({summary.get('avg_ttft_ms', 0):.1f}ms)"),
            ('平均 TPS', f"{summary.get('avg_tps', 0):.2f} tokens/s"),
            ('平均输出 Tokens', f"{summary.get('avg_tokens', 0):.0f}"),
            ('测试开始时间', summary.get('start_time', '')),
        ]
        
        for i, (label, value) in enumerate(stats, 4):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'A{i}'].border = self.BORDER
            ws[f'B{i}'].border = self.BORDER
        
        # 设置列宽
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 35
        
        # 添加说明
        ws['A13'] = '说明：'
        ws['A13'].font = Font(bold=True, color="666666")
        ws['A14'] = '• TTFT: Time To First Token，首个 Token 响应时间'
        ws['A15'] = '• TPS: Tokens Per Second，Token 生成速度'
        ws['A16'] = f'• 报告生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    
    def _create_model_comparison_sheet(self, wb: Workbook, model_stats: List[Dict]):
        """
        模型对比 Sheet
        
        Args:
            wb: Workbook 实例
            model_stats: 模型统计数据列表
        """
        ws = wb.create_sheet("模型对比")
        
        # 标题
        ws['A1'] = '模型性能对比表'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:H1')
        
        # 表头
        headers = ['排名', '模型名称', '测试次数', '成功数', '成功率', '平均 TTFT', '平均 TPS', '平均 Tokens']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER
        
        # 按性能排序（按平均 TPS 降序）
        sorted_stats = sorted(
            model_stats, 
            key=lambda x: x.get('avg_tps', 0), 
            reverse=True
        )
        
        # 数据行
        for row_idx, stat in enumerate(sorted_stats, 4):
            ws.cell(row=row_idx, column=1, value=row_idx - 3)  # 排名
            ws.cell(row=row_idx, column=2, value=stat.get('model_name', 'Unknown'))  # 模型名称
            ws.cell(row=row_idx, column=3, value=stat.get('count', 0))  # 测试次数
            ws.cell(row=row_idx, column=4, value=stat.get('success_count', 0))  # 成功数
            
            # 成功率
            count = stat.get('count', 0)
            success_count = stat.get('success_count', 0)
            success_rate = (success_count / count * 100) if count > 0 else 0
            ws.cell(row=row_idx, column=5, value=f"{success_rate:.1f}%")
            
            # 平均 TTFT
            avg_ttft = stat.get('avg_ttft', 0)
            ws.cell(row=row_idx, column=6, value=f"{avg_ttft:.3f}s")
            
            # 平均 TPS
            ws.cell(row=row_idx, column=7, value=f"{stat.get('avg_tps', 0):.2f}")
            
            # 平均 Tokens
            ws.cell(row=row_idx, column=8, value=f"{stat.get('avg_tokens', 0):.0f}")
            
            # 添加边框
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).border = self.BORDER
        
        # 设置列宽
        col_widths = [8, 25, 12, 12, 12, 15, 15, 15]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_detail_sheet(self, wb: Workbook, results: List[Dict]):
        """
        详细数据 Sheet
        
        Args:
            wb: Workbook 实例
            results: 测试结果列表
        """
        ws = wb.create_sheet("详细数据")
        
        # 标题
        ws['A1'] = f'测试详细记录 (共 {len(results)} 条)'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:J1')
        
        # 表头
        headers = ['序号', '时间戳', '模型', '测试用例', '轮次', 'TTFT(s)', 'TPS', 'Tokens', 'Think Tokens', '状态']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER
        
        # 数据行
        for row_idx, r in enumerate(results, 4):
            ws.cell(row=row_idx, column=1, value=row_idx - 3)  # 序号
            ws.cell(row=row_idx, column=2, value=r.get('timestamp', ''))  # 时间戳
            ws.cell(row=row_idx, column=3, value=r.get('model_name', ''))  # 模型
            ws.cell(row=row_idx, column=4, value=r.get('test_case_name', ''))  # 测试用例
            ws.cell(row=row_idx, column=5, value=r.get('round_number', 0))  # 轮次
            
            # TTFT
            ttft = r.get('ttft_seconds', 0)
            ws.cell(row=row_idx, column=6, value=round(ttft, 4))
            
            # TPS
            tps = r.get('tokens_per_second', 0)
            ws.cell(row=row_idx, column=7, value=round(tps, 2))
            
            # Tokens
            tokens = r.get('output_tokens', 0)
            ws.cell(row=row_idx, column=8, value=tokens)
            
            # Think Tokens（如果有）
            think_tokens = r.get('think_tokens', 0)
            ws.cell(row=row_idx, column=9, value=think_tokens)
            
            # 状态
            success = r.get('success', 0)
            status_cell = ws.cell(row=row_idx, column=10, value='成功' if success else '失败')
            
            # 状态颜色
            if success:
                status_cell.fill = self.SUCCESS_FILL
            else:
                status_cell.fill = self.ERROR_FILL
            
            # 添加边框
            for col in range(1, 11):
                ws.cell(row=row_idx, column=col).border = self.BORDER
        
        # 设置列宽
        col_widths = [8, 22, 20, 18, 8, 12, 10, 10, 12, 10]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_chart_data_sheet(self, wb: Workbook, chart_data: Dict):
        """
        图表数据 Sheet
        
        Args:
            wb: Workbook 实例
            chart_data: 图表数据
        """
        ws = wb.create_sheet("图表数据")
        
        # TTFT 趋势
        ws['A1'] = '性能趋势数据'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')
        
        # 获取模型列表
        models = list(chart_data.get('models', {}).keys())
        
        # TTFT 趋势表
        ws['A3'] = 'TTFT 趋势'
        ws['A3'].font = Font(bold=True)
        
        headers = ['轮次'] + models
        for col, header in enumerate(headers, 1):
            ws.cell(row=4, column=col, value=header)
            ws.cell(row=4, column=col).font = Font(bold=True)
            ws.cell(row=4, column=col).fill = self.HEADER_FILL
            ws.cell(row=4, column=col).font = self.HEADER_FONT
        
        trend_data = chart_data.get('ttft_trend', {})
        for row_idx, (round_num, values) in enumerate(trend_data.items(), 5):
            ws.cell(row=row_idx, column=1, value=f"R{round_num}")
            for col_idx, model in enumerate(models, 2):
                ws.cell(row=row_idx, column=col_idx, value=values.get(model, 0))
        
        # TPS 趋势表
        start_row = 5 + len(trend_data) + 2
        ws.cell(row=start_row, column=1, value='TPS 趋势')
        ws.cell(row=start_row, column=1).font = Font(bold=True)
        
        headers = ['轮次'] + models
        for col, header in enumerate(headers, 1):
            ws.cell(row=start_row + 1, column=col, value=header)
            ws.cell(row=start_row + 1, column=col).font = Font(bold=True)
            ws.cell(row=start_row + 1, column=col).fill = self.HEADER_FILL
            ws.cell(row=start_row + 1, column=col).font = self.HEADER_FONT
        
        tps_data = chart_data.get('tps_trend', {})
        for row_idx, (round_num, values) in enumerate(tps_data.items(), start_row + 2):
            ws.cell(row=row_idx, column=1, value=f"R{round_num}")
            for col_idx, model in enumerate(models, 2):
                ws.cell(row=row_idx, column=col_idx, value=values.get(model, 0))
        
        # 设置列宽
        ws.column_dimensions['A'].width = 10
        for col in range(2, len(models) + 3):
            ws.column_dimensions[get_column_letter(col)].width = 15


def export_to_excel(group_id: str) -> bytes:
    """
    快速导出 Excel 的便捷函数
    
    Args:
        group_id: 测试组 ID
        
    Returns:
        Excel 文件的字节数据
    """
    try:
        # 尝试多种导入方式以兼容不同运行环境
        try:
            from . import get_database as _get_db
            db = _get_db()
        except (ImportError, TypeError):
            try:
                from ..src.database import get_database
                db = get_database()
            except ImportError:
                from src.database import get_database
                db = get_database()
        
        # 获取汇总信息
        summary_data = db.get_group_summary(group_id)
        
        if not summary_data:
            raise ValueError(f"测试组 {group_id} 不存在")
        
        group_info = summary_data.get('group', {})
        model_stats = summary_data.get('model_stats', [])
        
        # 获取所有结果
        results = db.get_results(group_id)
        
        # 计算统计数据
        total = len(results)
        success = sum(1 for r in results if r.get('success'))
        failed = total - success
        success_rate = (success / total * 100) if total > 0 else 0
        
        # 计算平均值
        ttft_values = [r.get('ttft_seconds', 0) for r in results if r.get('ttft_seconds')]
        tps_values = [r.get('tokens_per_second', 0) for r in results if r.get('tokens_per_second')]
        token_values = [r.get('output_tokens', 0) for r in results if r.get('output_tokens')]
        
        avg_ttft = sum(ttft_values) / len(ttft_values) if ttft_values else 0
        avg_tps = sum(tps_values) / len(tps_values) if tps_values else 0
        avg_tokens = sum(token_values) / len(token_values) if token_values else 0
        
        summary = {
            'group_id': group_id,
            'start_time': group_info.get('start_time', ''),
            'total': total,
            'success': success,
            'failed': failed,
            'success_rate': success_rate,
            'avg_ttft': avg_ttft,
            'avg_ttft_ms': avg_ttft * 1000,
            'avg_tps': avg_tps,
            'avg_tokens': avg_tokens
        }
        
        # 格式化 model_stats
        formatted_stats = []
        for stat in model_stats:
            count = stat.get('total', 0)
            success_count = stat.get('success_count', 0)
            formatted_stats.append({
                'model_name': stat.get('model_name', 'Unknown'),
                'count': count,
                'success_count': success_count,
                'avg_ttft': stat.get('avg_ttft', 0),
                'avg_tps': stat.get('avg_tokens_per_second', 0),
                'avg_tokens': stat.get('total_output_tokens', 0) / count if count > 0 else 0
            })
        
        exporter = ExcelExporter()
        return exporter.export({
            'summary': summary,
            'results': results,
            'model_stats': formatted_stats,
            'chart_data': {}  # TODO: 如果需要图表数据，可以在这里生成
        })
        
    except ImportError as e:
        raise ImportError(f"无法导入必要的模块: {e}")
